import PyPDF2
import openpyxl

with open("out_utf8.txt", "w", encoding="utf-8") as out_f:
    out_f.write("--- PDF Content ---\n")
    try:
        with open("5COSC022W_Coursework Specification_Ref-Def(1).pdf", "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for i in range(len(reader.pages)):
                out_f.write(f"Page {i+1}:\n")
                out_f.write(reader.pages[i].extract_text() + "\n")
    except Exception as e:
        out_f.write(f"Error reading PDF: {e}\n")

    out_f.write("\n--- Excel Content ---\n")
    try:
        wb = openpyxl.load_workbook("5COSC022W_Coursework_Rubric_Ref_Def.xlsx")
        for sheet_name in wb.sheetnames:
            out_f.write(f"Sheet: {sheet_name}\n")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                out_f.write(f"{row}\n")
    except Exception as e:
        out_f.write(f"Error reading Excel: {e}\n")
